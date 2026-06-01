"""
Build the NLSS IV CHE audit workbook.

Output:
    6_output/main_output/audit_workbook.xlsx

The workbook reconciles the cleaned Stata dataset with Stata-produced CSV
outputs and records household-level traces for derived OOPG/OOP variables.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
import re
from typing import Iterable

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DATA_REL = Path("1_data") / "2_clean" / "catastrophic_health_exp.dta"
OFFICIAL_POVERTY = 20.27


def find_project_root() -> Path:
    cand = Path(__file__).resolve().parents[1]
    for _ in range(8):
        if (cand / DATA_REL).exists():
            return cand
        if cand == cand.parent:
            break
        cand = cand.parent
    raise FileNotFoundError(f"Could not locate {DATA_REL}")


PROJECT_ROOT = find_project_root()
DATA_PATH = PROJECT_ROOT / DATA_REL
OUT_DIR = PROJECT_ROOT / "6_output" / "main_output"
OUT_DIR.mkdir(parents=True, exist_ok=True)
OUT_PATH = OUT_DIR / "audit_workbook.xlsx"


def load_stata(path: Path) -> tuple[pd.DataFrame, dict[str, str]]:
    with pd.io.stata.StataReader(path) as reader:
        labels = reader.variable_labels()
        data = reader.read(convert_categoricals=False)
    return data, labels


def wmean(values: Iterable[float], weights: Iterable[float]) -> float:
    v = np.asarray(values, dtype=float)
    w = np.asarray(weights, dtype=float)
    mask = np.isfinite(v) & np.isfinite(w) & (w > 0)
    return float(np.sum(v[mask] * w[mask]) / np.sum(w[mask]))


def pct_headcount(flag: Iterable[float], weights: Iterable[float]) -> float:
    return 100.0 * wmean(flag, weights)


def read_csv_output(name: str) -> pd.DataFrame:
    path = OUT_DIR / name
    if path.exists():
        return pd.read_csv(path)
    return pd.DataFrame()


def derived_dictionary(labels: dict[str, str]) -> pd.DataFrame:
    rows = [
        ("oopg", "monthly NPR", "nominal", "HH", "30 days", "Section 8B communicable/injury OOP."),
        ("oop", "monthly NPR", "nominal", "HH", "mixed", "oopg plus annual NCD OOP divided by 12."),
        ("oopg_ann", "annual NPR", "nominal", "HH", "annualized 30 days", "oopg multiplied by 12."),
        ("oop_ann", "annual NPR", "nominal", "HH", "mixed annual", "oopg_ann plus annual NCD OOP."),
        ("oopg_pc_ann_real", "annual NPR", "real", "per_capita", "annualized 30 days", "(oopg_ann / hhsize) / paasche."),
        ("oop_pc_ann_real", "annual NPR", "real", "per_capita", "mixed annual", "(oop_ann / hhsize) / paasche."),
        ("oopg_real", "monthly NPR", "real", "HH", "30 days", "oopg / paasche."),
        ("oop_real", "monthly NPR", "real", "HH", "mixed monthly", "oop / paasche."),
        ("nonfood_spatial_index", "index", "nominalizing factor", "HH", "annual", "Domain nonfood spatial price index from the NLSS-IV methodological note."),
        ("nominal_pcexp_annual", "annual NPR", "nominal", "per_capita", "annual", "pcep_food * paasche + pcep_nonfood * nonfood_spatial_index."),
        ("nominal_pcexp_month", "monthly NPR", "nominal", "per_capita", "annual", "nominal_pcexp_annual / 12."),
        ("nominal_total_cons_annual", "annual NPR", "nominal", "HH", "annual", "nominal_pcexp_annual * hhsize."),
        ("nominal_total_cons_month", "monthly NPR", "nominal", "HH", "annual", "nominal_total_cons_annual / 12."),
        ("tot_real_hh_mo", "monthly NPR", "real", "HH", "annual", "pcep * hhsize / 12."),
        ("nf_real_hh_mo", "monthly NPR", "real", "HH", "annual", "pcep_nonfood * hhsize / 12."),
        ("totexp_nom_oopgadd_mo", "monthly NPR", "nominal", "HH", "annual plus 30 days", "nominal_total_cons_month + oopg."),
        ("ctp_real_oopgadd_hh_mo", "monthly NPR", "real", "HH", "annual plus 30 days", "nf_real_hh_mo + oopg_real."),
        ("totexp_real_hh_mo", "monthly NPR", "real", "HH", "audit annual plus health", "Audit: tot_real_hh_mo + oop_real."),
        ("ctp_real_hh_mo", "monthly NPR", "real", "HH", "audit annual plus health", "Audit: nf_real_hh_mo + oop_real."),
        ("pre_oopg", "annual NPR", "real", "per_capita", "annualized 30 days", "pcep + oopg_pc_ann_real."),
        ("pre_oop", "annual NPR", "real", "per_capita", "mixed annual", "pcep + oop_pc_ann_real."),
    ]

    for name in ["oopg_sh_tot", "oop_sh_tot", "oopg_sh_nf", "oop_sh_nf"]:
        denom = "nominal total consumption plus OOPG" if name.endswith("_tot") else "real nonfood consumption plus OOPG"
        real_flag = "nominal/nominal" if name.endswith("_tot") else "real/real"
        rows.append((name, "share", real_flag, "HH", "monthly", f"Health-payment numerator divided by {denom}."))

    for name in ["oopg_sh_tot_nlss", "oop_sh_tot_nlss", "oopg_sh_nf_nlss", "oop_sh_nf_nlss"]:
        denom = "NLSS total consumption excluding health" if "_tot_" in name else "NLSS nonfood consumption excluding health"
        rows.append((name, "share", "real/real", "HH", "monthly", f"Health-payment numerator divided by {denom}."))

    for scope in ["oopg", "oop"]:
        for suffix, thresh in [("tot10", "10% total"), ("tot20", "20% total"), ("nf25", "25% capacity to pay"), ("nf40", "40% capacity to pay")]:
            rows.append((f"che_{scope}_{suffix}", "0/1", "mixed", "HH", "monthly", f"Indicator that {scope} exceeds {thresh} using the primary OOPG-addback denominator."))
            rows.append((f"che_{scope}_{suffix}_nlss", "0/1", "real/real", "HH", "monthly", f"Indicator that {scope} exceeds {thresh} using NLSS denominator."))
            rows.append((f"over_{scope}_{suffix}", "share points", "mixed", "HH", "monthly", f"Positive excess above {thresh} using the primary OOPG-addback denominator."))

    out = []
    for name, units, real, scope, window, formula in rows:
        out.append(
            {
                "name": name,
                "label": labels.get(name, ""),
                "units": units,
                "nominal_or_real": real,
                "scope": scope,
                "recall_window": window,
                "formula_in_plain_language": formula,
                "source_script": "2_prep/1_catastrophic.do",
            }
        )
    return pd.DataFrame(out).drop_duplicates("name")


def edge_case_trace(df: pd.DataFrame) -> pd.DataFrame:
    base_cols = [
        "psu_number",
        "hh_number",
        "hhsize",
        "paasche",
        "hh_comm_total_30d",
        "hh_ncd_total_annual",
        "pcep",
        "pcep_food",
        "pcep_nonfood",
        "oopg",
        "oop",
        "oopg_real",
        "oop_real",
        "nominal_total_cons_month",
        "tot_real_hh_mo",
        "nf_real_hh_mo",
        "totexp_nom_oopgadd_mo",
        "ctp_real_oopgadd_hh_mo",
        "totexp_real_hh_mo",
        "ctp_real_hh_mo",
        "oopg_sh_tot",
        "oop_sh_tot",
        "oopg_sh_nf",
        "oop_sh_nf",
        "che_oopg_tot10",
        "che_oop_tot10",
        "che_oopg_tot20",
        "che_oop_tot20",
        "che_oopg_nf25",
        "che_oop_nf25",
        "che_oopg_nf40",
        "che_oop_nf40",
    ]

    frames = []
    for label, data in [
        ("lowest_pcep", df.nsmallest(5, "pcep")),
        ("highest_pcep", df.nlargest(5, "pcep")),
        ("highest_oop", df.nlargest(5, "oop")),
    ]:
        tmp = data.copy()
        tmp.insert(0, "selection_group", label)
        frames.append(tmp)

    cross = df[(df["che_oopg_tot10"] == 0) & (df["che_oop_tot10"] == 1)].copy()
    if len(cross) < 5:
        cross = df.assign(_near=(df["oop_sh_tot"] - 0.10).abs()).nsmallest(5, "_near")
    else:
        cross = cross.nlargest(5, "oop_sh_tot")
    cross.insert(0, "selection_group", "oop_only_or_nearest_tot10")
    frames.append(cross)

    out = pd.concat(frames, ignore_index=True)
    out = out.drop_duplicates(["psu_number", "hh_number"], keep="first")
    if len(out) < 20:
        fill = df.nlargest(20, "oop").copy()
        fill.insert(0, "selection_group", "fill_highest_oop")
        out = pd.concat([out, fill], ignore_index=True).drop_duplicates(["psu_number", "hh_number"], keep="first")
    return out[["selection_group", *base_cols]].head(20)


def python_headlines(df: pd.DataFrame) -> dict[str, float]:
    ind_wt = df["ind_wt"].astype(float)
    out: dict[str, float] = {
        "N households": float(len(df)),
        "pcep < pline headcount (ind_wt)": pct_headcount(df["pcep"] < df["pline"], ind_wt),
        "Mean pcep": wmean(df["pcep"], ind_wt),
        "Mean pcep_nonfood": wmean(df["pcep_nonfood"], ind_wt),
        "Mean pcep_food": wmean(df["pcep_food"], ind_wt),
    }

    for conv in ["primary", "nlss"]:
        suffix = "" if conv == "primary" else "_nlss"
        for scope in ["oopg", "oop"]:
            for thresh in ["tot10", "tot20", "nf25", "nf40"]:
                var = f"che_{scope}_{thresh}{suffix}"
                out[f"CHE {scope} {thresh} {conv} prevalence (ind_wt)"] = pct_headcount(df[var] == 1, ind_wt)

    for scope in ["oopg", "oop"]:
        pre = df[f"pre_{scope}"].astype(float)
        post = df["pcep"].astype(float)
        poor_pre = pre < df["pline"]
        poor_post = post < df["pline"]
        pushed = (~poor_pre) & poor_post
        pre_h = pct_headcount(poor_pre, ind_wt)
        post_h = pct_headcount(poor_post, ind_wt)
        out[f"{scope} pre-OOP poverty headcount"] = pre_h
        out[f"{scope} post-OOP poverty headcount"] = post_h
        out[f"{scope} OOP-induced impoverishment pp"] = post_h - pre_h
        out[f"{scope} people pushed into poverty"] = float(ind_wt[pushed].sum())
    return out


def stata_headlines() -> dict[str, float]:
    out: dict[str, float] = {}

    for log_path in [PROJECT_ROOT / "0_master.log", PROJECT_ROOT / "4_log" / "3_poverty_impact.log"]:
        if log_path.exists():
            text = log_path.read_text(errors="ignore")
            hit = re.search(r"Observations:\s*([0-9,]+)", text)
            if not hit:
                hit = re.search(r"Total households:\s*([0-9,]+)", text)
            if hit:
                out["N households"] = float(hit.group(1).replace(",", ""))
                break

    desc = read_csv_output("descriptive_means.csv")
    if not desc.empty:
        for var in ["pcep", "pcep_nonfood", "pcep_food"]:
            hit = desc[(desc["variable"] == var) & (desc["weight"] == "ind_wt")]
            if not hit.empty:
                out[f"Mean {var}"] = float(hit.iloc[0]["estimate"])

    prevalence = read_csv_output("che_prevalence_all.csv")
    if not prevalence.empty:
        hit = prevalence[prevalence["weight"] == "ind_wt"]
        for _, row in hit.iterrows():
            thresh = "tot10" if row["denominator"] == "total" and abs(row["threshold"] - 0.10) < 1e-9 else ""
            if row["denominator"] == "total" and abs(row["threshold"] - 0.20) < 1e-9:
                thresh = "tot20"
            if row["denominator"] == "nonfood" and abs(row["threshold"] - 0.25) < 1e-9:
                thresh = "nf25"
            if row["denominator"] == "nonfood" and abs(row["threshold"] - 0.40) < 1e-9:
                thresh = "nf40"
            out[f"CHE {row['scope']} {thresh} {row['convention']} prevalence (ind_wt)"] = 100.0 * float(row["prevalence"])

    pov = read_csv_output("poverty_impact.csv")
    if not pov.empty:
        for scope in ["oopg", "oop"]:
            hit = pov[(pov["scope"] == scope) & (pov["metric"] == "poverty_headcount")]
            if not hit.empty:
                row = hit.iloc[0]
                out[f"{scope} pre-OOP poverty headcount"] = 100.0 * float(row["pre_estimate"])
                out[f"{scope} post-OOP poverty headcount"] = 100.0 * float(row["post_estimate"])
                out[f"{scope} OOP-induced impoverishment pp"] = 100.0 * float(row["difference"])
                out[f"{scope} people pushed into poverty"] = float(row["people_pushed"])
                out["pcep < pline headcount (ind_wt)"] = 100.0 * float(row["post_estimate"])

    out.setdefault("N households", np.nan)
    return out


def reconciliation_table(df: pd.DataFrame) -> pd.DataFrame:
    py = python_headlines(df)
    st = stata_headlines()
    rows = []
    for name, py_val in py.items():
        st_val = st.get(name, np.nan)
        abs_diff = abs(st_val - py_val) if np.isfinite(st_val) else np.nan
        pct_diff = abs_diff / abs(st_val) * 100 if np.isfinite(st_val) and st_val != 0 else np.nan
        rows.append(
            {
                "statistic_name": name,
                "stata_value": st_val,
                "python_value": py_val,
                "abs_diff": abs_diff,
                "pct_diff": pct_diff,
            }
        )
    return pd.DataFrame(rows)


def sanity_checks(df: pd.DataFrame) -> pd.DataFrame:
    checks = []

    def add(name: str, passed: bool | None, detail: str = "") -> None:
        status = "PENDING" if passed is None else ("PASS" if passed else "FAIL")
        checks.append({"check": name, "status": status, "detail": detail})

    ind_wt = df["ind_wt"].astype(float)
    poor = pct_headcount(df["pcep"] < df["pline"], ind_wt)

    add("N households = 9,600", len(df) == 9600, f"N={len(df):,}")
    add("pcep < pline (ind_wt) = 20.27% +/- 0.05 pp", abs(poor - OFFICIAL_POVERTY) <= 0.05, f"{poor:.4f}%")
    add("max |pcep - (pcep_food + pcep_nonfood)| < 1 NPR", float((df["pcep"] - (df["pcep_food"] + df["pcep_nonfood"])).abs().max()) < 1)
    add("oop >= oopg for all households", bool((df["oop"] + 1e-9 >= df["oopg"]).all()))

    dominance = True
    for thresh in ["tot10", "tot20", "nf25", "nf40"]:
        dominance &= bool((df[f"che_oop_{thresh}"] >= df[f"che_oopg_{thresh}"]).all())
        dominance &= bool((df[f"che_oop_{thresh}_nlss"] >= df[f"che_oopg_{thresh}_nlss"]).all())
    add("OOP CHE flags dominate OOPG flags", dominance)

    mono = True
    for scope in ["oopg", "oop"]:
        mono &= bool((df[f"che_{scope}_tot20"] <= df[f"che_{scope}_tot10"]).all())
        mono &= bool((df[f"che_{scope}_nf40"] <= df[f"che_{scope}_nf25"]).all())
    add("CHE threshold monotonicity", mono)

    over_ok = True
    for scope in ["oopg", "oop"]:
        for thresh in ["tot10", "tot20", "nf25", "nf40"]:
            over = df[f"over_{scope}_{thresh}"]
            che = df[f"che_{scope}_{thresh}"]
            over_ok &= bool(((over == 0) == (che == 0)).all())
    add("over_X = 0 iff che_X = 0", over_ok)

    equality = (df["oop_pc_ann_real"].abs() < 1e-9)
    pre_ok = bool((df["pre_oop"] + 1e-9 >= df["pcep"]).all())
    pre_eq_ok = bool(((df["pre_oop"] - df["pcep"]).abs() < 1e-9).eq(equality).all())
    add("pre_oop >= pcep; equality iff oop_pc_ann_real = 0", pre_ok and pre_eq_ok)
    add("ctp_real_oopgadd_hh_mo > 0 for all households with hhsize > 0", bool((df.loc[df["hhsize"] > 0, "ctp_real_oopgadd_hh_mo"] > 0).all()))
    add("totexp_nom_oopgadd_mo > 0 for all households", bool((df["totexp_nom_oopgadd_mo"] > 0).all()))

    tot_eq = df["oop"].abs() < 1e-9
    tot_ok = bool((df["totexp_real_hh_mo"] + 1e-9 >= df["tot_real_hh_mo"]).all())
    tot_eq_ok = bool(((df["totexp_real_hh_mo"] - df["tot_real_hh_mo"]).abs() < 1e-9).eq(tot_eq).all())
    add("audit totexp_real_hh_mo >= tot_real_hh_mo; equality iff oop = 0", tot_ok and tot_eq_ok)

    add("reconstructed nominal total consumption is positive", bool((df["nominal_total_cons_month"] > 0).all()))
    nom_eq = df["oopg"].abs() < 1e-9
    nom_ok = bool((df["totexp_nom_oopgadd_mo"] + 1e-9 >= df["nominal_total_cons_month"]).all())
    nom_eq_ok = bool(((df["totexp_nom_oopgadd_mo"] - df["nominal_total_cons_month"]).abs() < 1e-9).eq(nom_eq).all())
    add("primary nominal total denominator >= reconstructed nominal consumption; equality iff oopg = 0", nom_ok and nom_eq_ok)

    cmp = df.loc[df["oop"] > 0, "oopg_sh_nf"] <= df.loc[df["oop"] > 0, "oopg_sh_nf_nlss"] + 1e-12
    mean_primary = wmean(df["oopg_sh_nf"], df["hhs_wt"])
    mean_nlss = wmean(df["oopg_sh_nf_nlss"], df["hhs_wt"])
    add("oopg_sh_nf <= oopg_sh_nf_nlss; strict at population mean", bool(cmp.all() and mean_primary < mean_nlss))

    add("WHO NHA OOP per capita comparison", None, "pending manual entry; external NHA value not fetched by this script")
    return pd.DataFrame(checks)


def prevalence_table(df: pd.DataFrame) -> pd.DataFrame:
    prev = read_csv_output("che_prevalence_all.csv")
    if prev.empty:
        rows = []
        for convention in ["primary", "nlss"]:
            suffix = "" if convention == "primary" else "_nlss"
            for scope in ["oopg", "oop"]:
                for denominator, thresholds in [("total", [("tot10", 0.10), ("tot20", 0.20)]), ("nonfood", [("nf25", 0.25), ("nf40", 0.40)])]:
                    for code, threshold in thresholds:
                        var = f"che_{scope}_{code}{suffix}"
                        for wt in ["hhs_wt", "ind_wt"]:
                            p = wmean(df[var], df[wt])
                            rows.append({"weight": wt, "scope": scope, "denominator": denominator, "threshold": threshold, "convention": convention, "variable": var, "prevalence": p, "se": np.nan, "ci_low": np.nan, "ci_high": np.nan})
        prev = pd.DataFrame(rows)

    wide_rows = []
    keys = ["scope", "denominator", "threshold", "convention", "variable"]
    for key, grp in prev.groupby(keys, dropna=False):
        row = dict(zip(keys, key))
        for wt in ["hhs_wt", "ind_wt"]:
            hit = grp[grp["weight"] == wt]
            if not hit.empty:
                h = hit.iloc[0]
                prefix = "hh" if wt == "hhs_wt" else "ind"
                row[f"{prefix}_prevalence_pct"] = 100.0 * h["prevalence"]
                row[f"{prefix}_ci_low_pct"] = 100.0 * h["ci_low"]
                row[f"{prefix}_ci_high_pct"] = 100.0 * h["ci_high"]
        wide_rows.append(row)
    return pd.DataFrame(wide_rows).sort_values(["scope", "denominator", "threshold", "convention"])


def impoverishment_summary(df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    ind_wt = df["ind_wt"].astype(float)
    post = df["pcep"].astype(float)
    poor_post = post < df["pline"]
    food_post = post < df["fpline"]

    for scope in ["oopg", "oop"]:
        pre = df[f"pre_{scope}"].astype(float)
        poor_pre = pre < df["pline"]
        food_pre = pre < df["fpline"]
        pushed = (~poor_pre) & poor_post
        rows.append(
            {
                "scope": scope,
                "pre_OOP_headcount_pct": pct_headcount(poor_pre, ind_wt),
                "post_OOP_headcount_pct": pct_headcount(poor_post, ind_wt),
                "impoverishment_pp": pct_headcount(poor_post, ind_wt) - pct_headcount(poor_pre, ind_wt),
                "absolute_people_pushed": float(ind_wt[pushed].sum()),
                "households_crossing_threshold": int(pushed.sum()),
                "pre_OOP_food_poverty_headcount_pct": pct_headcount(food_pre, ind_wt),
                "post_OOP_food_poverty_headcount_pct": pct_headcount(food_post, ind_wt),
            }
        )
    return pd.DataFrame(rows)


def readme_sheet() -> pd.DataFrame:
    purpose = (
        "Audit workbook for the Nepal NLSS IV catastrophic health expenditure "
        "pipeline. It documents derived variables, traces edge-case households, "
        "reconciles Stata and Python headline estimates, and records sanity "
        "checks for OOPG and OOP analysis."
    )
    sheets = [
        ("README", "Purpose, contents, generation metadata."),
        ("variable_dictionary", "Definitions, units, scope, and formulas for derived variables."),
        ("derived_variable_trace", "Twenty edge-case households for checking formulas by hand."),
        ("stata_vs_python_reconciliation", "Headline Stata output values compared with Python calculations."),
        ("sanity_checks", "Pass/fail checks for the cleaned dataset and derived variables."),
        ("prevalence_table", "CHE prevalence under primary and NLSS-style denominators."),
        ("impoverishment_summary", "Pre/post poverty and food-poverty summaries by scope."),
    ]
    rows = [
        {"field": "purpose", "value": purpose},
        {"field": "generation_date", "value": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
        {"field": "NLSS_cycle", "value": "NLSS IV (2022/23)"},
        {"field": "project_commit_hash", "value": "not recorded; git calls avoided per project guidance"},
    ]
    rows.extend({"field": f"sheet: {name}", "value": desc} for name, desc in sheets)
    return pd.DataFrame(rows)


def apply_formatting(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    fail_fill = PatternFill("solid", fgColor="FFC7CE")
    pending_fill = PatternFill("solid", fgColor="FFF2CC")

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True)
        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            max_len = max(len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, min(ws.max_row, 60) + 1))
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 45)

    if "sanity_checks" in wb.sheetnames:
        ws = wb["sanity_checks"]
        status_col = None
        for cell in ws[1]:
            if cell.value == "status":
                status_col = cell.column
                break
        if status_col:
            letter = get_column_letter(status_col)
            ws.conditional_formatting.add(
                f"{letter}2:{letter}{ws.max_row}",
                CellIsRule(operator="equal", formula=['"FAIL"'], fill=fail_fill),
            )
            ws.conditional_formatting.add(
                f"{letter}2:{letter}{ws.max_row}",
                CellIsRule(operator="equal", formula=['"PENDING"'], fill=pending_fill),
            )

    if "stata_vs_python_reconciliation" in wb.sheetnames:
        ws = wb["stata_vs_python_reconciliation"]
        for col_name in ["abs_diff", "pct_diff"]:
            col = None
            for cell in ws[1]:
                if cell.value == col_name:
                    col = cell.column
                    break
            if col:
                letter = get_column_letter(col)
                threshold = "0.5" if col_name == "abs_diff" else "1"
                ws.conditional_formatting.add(
                    f"{letter}2:{letter}{ws.max_row}",
                    CellIsRule(operator="greaterThan", formula=[threshold], fill=fail_fill),
                )

    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, float):
                    header = ws.cell(row=1, column=cell.column).value or ""
                    if "pct" in str(header).lower() or "headcount" in str(header).lower() or "prevalence" in str(header).lower() or "pp" in str(header).lower():
                        cell.number_format = "0.00"
                    elif "people" in str(header).lower() or "household" in str(header).lower() or "count" in str(header).lower():
                        cell.number_format = '#,##0'
                    else:
                        cell.number_format = '#,##0.00'

    wb.save(path)


def main() -> None:
    df, labels = load_stata(DATA_PATH)
    assert len(df) == 9600, f"Expected 9,600 households, found {len(df):,}"
    assert abs(df["pcep"].astype(float) - (df["pcep_food"].astype(float) + df["pcep_nonfood"].astype(float))).max() < 1
    assert (df["oop"].astype(float) + 1e-9 >= df["oopg"].astype(float)).all()
    assert (df["ctp_real_oopgadd_hh_mo"].astype(float) > 0).all()
    assert (df["totexp_nom_oopgadd_mo"].astype(float) > 0).all()
    assert (df["pre_oop"].astype(float) + 1e-9 >= df["pcep"].astype(float)).all()
    poor = pct_headcount(df["pcep"] < df["pline"], df["ind_wt"])
    assert abs(poor - OFFICIAL_POVERTY) <= 0.05, f"Official poverty check failed: {poor:.4f}%"

    sheets = {
        "README": readme_sheet(),
        "variable_dictionary": derived_dictionary(labels),
        "derived_variable_trace": edge_case_trace(df),
        "stata_vs_python_reconciliation": reconciliation_table(df),
        "sanity_checks": sanity_checks(df),
        "prevalence_table": prevalence_table(df),
        "impoverishment_summary": impoverishment_summary(df),
    }

    with pd.ExcelWriter(OUT_PATH, engine="openpyxl") as writer:
        for name, table in sheets.items():
            table.to_excel(writer, sheet_name=name, index=False)

    apply_formatting(OUT_PATH)
    print(f"Saved: {OUT_PATH}")


if __name__ == "__main__":
    main()
